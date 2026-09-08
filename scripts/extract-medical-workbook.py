"""Extract only authorized medical product formulas and numeric reference tables.

Does not edit/export the source workbook. Usage: python script.py SOURCE.xlsx
"""
import hashlib
import json
import sys
import warnings
from pathlib import Path
import openpyxl

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
source = Path(sys.argv[1])
digest = hashlib.sha256(source.read_bytes()).hexdigest()
formulas = openpyxl.load_workbook(source, read_only=True, data_only=False)
cached = openpyxl.load_workbook(source, read_only=True, data_only=True)
result = {'version': '2026-08-03', 'sourceSha256': digest, 'sheets': {}}
for name, end in [('CIM3 BCIM3', 150), ('CIE3', 146), ('CIP2', 149)]:
    cells = {}
    for row in formulas[name].iter_rows(max_row=end, max_col=56):
        for cell in row:
            v = cell.value
            if cell.data_type == 'f' and 'HYPERLINK' not in v:
                cells[cell.coordinate] = v
            elif isinstance(v, (int, float, bool)):
                cells[cell.coordinate] = v
    # All raw scenario strings are explicit adapter inputs, never author identity.
    result['sheets'][name] = cells
for name in ['dataCIM', 'dataCIE', 'dataCIP']:
    result['sheets'][name] = {c.coordinate: c.value for row in cached[name].iter_rows(max_row=108, max_col=341) for c in row if isinstance(c.value, (int, float, bool)) or getattr(c, 'column', 0) == 1 and c.value == 'ANB'}
result['sheets']['Notes'] = {a: cached['Notes'][a].value for a in ['B12', 'B13']}
assert hashlib.sha256(source.read_bytes()).hexdigest() == digest
destination = Path(__file__).resolve().parents[1] / 'src/lib/medicalWorkbook.json'
destination.write_text(json.dumps(result, ensure_ascii=False, separators=(',', ':')) + '\n')
print(f'Extracted {sum(len(s) for s in result["sheets"].values())} cells; source unchanged: {digest}')
