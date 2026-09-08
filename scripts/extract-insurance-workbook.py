"""Read-only extraction of the user-supplied calculator; never saves the workbook.

Usage: python scripts/extract-insurance-workbook.py /path/to/workbook.xlsx
Generated data deliberately excludes author identities, hyperlinks and other products.
"""
import json
import sys
from pathlib import Path
import openpyxl

source = Path(sys.argv[1])
formulas = openpyxl.load_workbook(source, read_only=True, data_only=False)
cached = openpyxl.load_workbook(source, read_only=True, data_only=True)
destination = Path(__file__).resolve().parents[1] / 'src/lib/insuranceWorkbook.json'
result = {'version': '2026-08-03', 'sheets': {}, 'baseline': {}}
for name, last, guard in [('TRST', 164, 'EI63'), ('PRMESP', 165, 'EM64')]:
    cells = {}
    for row in formulas[name].iter_rows(max_row=last):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            # Formula cells are evaluated on demand; captions are supplied by the UI.
            if cell.data_type == 'f' and 'HYPERLINK' not in v:
                cells[cell.coordinate] = v
            elif isinstance(v, (int, float, bool)):
                cells[cell.coordinate] = v
            elif cell.coordinate in ['C10', 'C12', 'C14', 'F12', 'F13', 'F14', 'H12', 'H13', 'H14', 'H62', 'H63']:
                cells[cell.coordinate] = v
    cells[guard] = 'JoJo'
    cells['H5'] = 'JoJo'
    cells['B178' if name == 'TRST' else 'B168'] = 'chi'
    result['sheets'][name] = cells
    start = 64 if name == 'TRST' else 65
    result['baseline'][name] = {
        c.coordinate: c.value for row in cached[name].iter_rows(min_row=start, max_row=last, max_col=12)
        for c in row if isinstance(c.value, (int, float))
    }
    for a in ['C16', 'C17', 'C18', 'F16', 'F17', 'F18', 'C59' if name == 'TRST' else 'C60']:
        v = cached[name][a].value
        if isinstance(v, (int, float)):
            result['baseline'][name][a] = v
# These tables contain no scenario inputs. Retain only columns referenced by the two
# product formula families, plus lookup keys. Cached values are the workbook's data.
for name, first, last in [('data3', 243, 244), ('data4', 201, 248)]:
    result['sheets'][name] = {
        c.coordinate: c.value for row in cached[name] for c in row
        if isinstance(c.value, (int, float, str, bool)) and (c.column == 1 or first <= c.column <= last)
    }
result['sheets']['Notes'] = {a: cached['Notes'][a].value for a in ['B6', 'B10', 'B12', 'B13', 'B14']}
destination.write_text(json.dumps(result, ensure_ascii=False, separators=(',', ':')) + '\n')
print(f'Extracted {sum(len(v) for v in result["sheets"].values())} cells to {destination.name}')
